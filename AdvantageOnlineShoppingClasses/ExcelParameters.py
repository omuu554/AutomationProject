from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelParemters:
    def __init__(self):
        self.workbook = load_workbook(filename="Parameters.xlsx")
        self.Parameters = self.workbook.active
        self.NumberToTestDict = {"1":"C", "2":"D", "3":"E", "4":"F", "5":"G", "6":"H", "7":"I", "8":"J", "9":"K", "10":"L"}

    def Get_TestParameters_Dict(self,TestNumber:int):
        TestNumber = TestNumber + 1
        ParametersDict = {}
        info = {}
        for row in self.Parameters.iter_rows(min_row=2, max_row=31, min_col=1, max_col=12, values_only=True):
            if (row[0] is not None):
                currentObject = row[0]
                info = {}
            info[row[1]] = row[TestNumber]
            ParametersDict[currentObject] = info

        return ParametersDict

    def Edit_CellByTestNumber(self,index:int, TestResult:str):
        index = str(index)
        TestCol = self.NumberToTestDict[index]
        self.Parameters[f"{TestCol}32"].value = TestResult




    def Close_Workbook(self):
        self.workbook.close()

    def Save_Workbook(self):
        self.workbook.save("ParametersWithResults.xlsx")










excelSheet = ExcelParemters()
print(excelSheet.Parameters["C7"].value)








